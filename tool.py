#!/usr/bin/env python3
"""
Hoa Don Dien Tu Auto-Fill Tool
Reads an Excel file, auto-fills the public invoice lookup form at
https://hoadondientu.gdt.gov.vn/lookup, and writes a results Excel file
with the original data plus an embedded screenshot in the last column.

Usage:
    python tool.py invoices.xlsx
    python tool.py invoices.xlsx --auto-captcha   # Claude AI solves CAPTCHA
    python tool.py invoices.xlsx --headless        # no visible browser window

Excel columns (case-insensitive):
    nbmst     - MST người bán (Tax code, 10 or 14 digits) [required]
    lhdon     - Loại hóa đơn (GTGT/HDBH/HDK/...) [default: GTGT]
    khhdon    - Ký hiệu hóa đơn [required]
    shdon     - Số hóa đơn (digits only) [required]
    tgtthue   - Tổng tiền thuế [optional]
    tgtttbso  - Tổng tiền thanh toán [required]
"""

import argparse
import base64
import io
import os
import random
import sys
import time
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
from playwright.sync_api import sync_playwright, Page, Locator

# ── Config ────────────────────────────────────────────────────────────────────

OUTPUT_DIR = Path("output")
LOOKUP_URL = "https://hoadondientu.gdt.gov.vn/"
MAX_CAPTCHA_RETRIES = 10

# Thumbnail size embedded in the results Excel
THUMB_W = 1920   # pixels wide
THUMB_H = 1080   # pixels tall (max — aspect ratio is preserved)

INVOICE_TYPE_MAP = {
    "GTGT":    "giá trị gia tăng",
    "HDBH":    "Hóa đơn bán hàng",
    "HBBTSC":  "bán tài sản công",
    "HDDTQG":  "dự trữ quốc gia",
    "HDK":     "Hóa đơn khác",
    "CT":      "phát hành, sử dụng và quản lý như hóa đơn",
    "GTGTMTT": "giá trị gia tăng khởi tạo từ máy tính tiền",
    "HDBHMTT": "bán hàng khởi tạo từ máy tính tiền",
}

# ── CAPTCHA helpers ───────────────────────────────────────────────────────────

def get_captcha_answer(page: Page, auto_solve: bool) -> str:
    img = page.locator(
        "img[src*='captcha'], img[src*='Captcha'], "
        ".ant-form-item:has-text('Mã captcha') img, "
        "img[alt*='captcha']"
    ).first
    try:
        img_bytes = img.screenshot()
    except Exception:
        img_bytes = None

    if img_bytes:
        Path("captcha_last.png").write_bytes(img_bytes)
    print("    [CAPTCHA] captcha_last.png saved — check browser window.")
    return input("    [CAPTCHA] Type the CAPTCHA: ").strip()


# ── Form-fill helpers ─────────────────────────────────────────────────────────

def _clear_fill(loc: Locator, value: str) -> None:
    loc.click()
    loc.press("Control+a")
    loc.press("Backspace")
    loc.fill(value)


def _fill_number(page: Page, label_text: str, value, field_id: str = "") -> bool:
    if value is None or str(value).strip() in ("", "nan"):
        return False
    raw = str(value).replace(",", "").replace(" ", "").strip()
    try:
        raw = str(int(float(raw)))
    except ValueError:
        return False

    inp = None
    # Try by ID first (most reliable)
    if field_id:
        loc = page.locator(f"input#{field_id}")
        if loc.count():
            inp = loc.first

    # Fallback: label text match
    if inp is None:
        item = page.locator(".ant-form-item").filter(has_text=label_text)
        loc = item.locator("input").first
        if loc.count():
            inp = loc

    if inp is None:
        print(f"    [WARN] Could not find input for '{label_text}'")
        return False

    inp.click()
    time.sleep(0.2)
    # Select all (Meta for macOS, Control for Linux/Windows)
    inp.press("Meta+a")
    inp.press("Backspace")
    inp.type(raw, delay=50)
    # Blur to trigger Ant Design value change
    inp.press("Tab")
    return True


def _select_invoice_type(page: Page, lhdon: str) -> None:
    lhdon = lhdon.upper().strip() if lhdon else "GTGT"
    label_text = INVOICE_TYPE_MAP.get(lhdon, INVOICE_TYPE_MAP["GTGT"])
    trigger = (
        page.locator(".ant-form-item")
        .filter(has_text="Loại hóa đơn")
        .locator(".ant-select-selector")
        .first
    )
    trigger.click()
    page.wait_for_selector(".ant-select-dropdown:visible", timeout=5000)
    option = page.locator(
        f'.ant-select-item-option[title*="{label_text}"],'
        f'.ant-select-item-option:has-text("{label_text}")'
    ).first
    if option.count():
        option.click()
    else:
        page.locator(".ant-select-item-option").first.click()


# ── Core browser logic ────────────────────────────────────────────────────────

def fill_form(page: Page, row: dict, auto_solve: bool, captcha_fn=None) -> None:
    page.wait_for_selector(".ant-form", timeout=20000)

    # ── Dismiss any announcement / popup modal that covers the form ───────
    try:
        # Try the close button (X) on the modal
        close_btn = page.locator(
            ".ant-modal-close, "
            ".ant-modal-wrap button.ant-modal-close, "
            "button[aria-label='Close']"
        ).first
        if close_btn.count() and close_btn.is_visible():
            close_btn.click(timeout=3000)
            page.wait_for_timeout(500)
    except Exception:
        pass

    # If a modal is still blocking, press Escape to dismiss it
    try:
        modal = page.locator(".ant-modal-wrap").first
        if modal.count() and modal.is_visible():
            page.keyboard.press("Escape")
            page.wait_for_timeout(500)
    except Exception:
        pass

    mst = str(row.get("nbmst", "")).strip()
    if mst:
        inp = page.locator(
            'input[id*="nbmst"], input[placeholder*="MST"]'
        ).or_(
            page.locator(".ant-form-item").filter(has_text="MST người bán").locator("input")
        ).first
        _clear_fill(inp, mst)

    lhdon = str(row.get("lhdon", "GTGT"))
    try:
        _select_invoice_type(page, lhdon)
    except Exception as exc:
        print(f"    [WARN] Invoice type select failed: {exc}")

    khhdon = str(row.get("khhdon", "")).strip()
    if khhdon:
        inp = page.locator(
            'input[id*="khhdon"], input[placeholder*="Ký hiệu"]'
        ).or_(
            page.locator(".ant-form-item").filter(has_text="Ký hiệu hóa đơn").locator("input")
        ).first
        _clear_fill(inp, khhdon)

    shdon = str(row.get("shdon", "")).strip()
    if shdon:
        inp = page.locator(
            'input[id*="shdon"], input[placeholder*="Số hóa đơn"]'
        ).or_(
            page.locator(".ant-form-item").filter(has_text="Số hóa đơn").locator("input")
        ).first
        _clear_fill(inp, shdon)

    _fill_number(page, "Tổng tiền thuế", row.get("tgtthue"), field_id="tgtthue")
    _fill_number(page, "Tổng tiền thanh toán", row.get("tgtttbso"), field_id="tgtttbso")

    solver = captcha_fn or get_captcha_answer
    captcha_answer = solver(page, auto_solve)
    captcha_inp = page.locator(
        'input[id*="captcha"], input[placeholder*="captcha"], '
        'input[placeholder*="Captcha"], input[placeholder*="mã"]'
    ).or_(
        page.locator(".ant-form-item").filter(has_text="Nhập mã captcha").locator("input")
    ).first
    _clear_fill(captcha_inp, captcha_answer)

    page.locator(
        'button[type="submit"]:has-text("Tìm kiếm"), button:has-text("Tìm kiếm")'
    ).first.click()
    page.wait_for_load_state("networkidle", timeout=20000)
    time.sleep(0.8)


# ── Screenshot embedding ──────────────────────────────────────────────────────

def _make_thumb(screenshot_path: str) -> tuple[io.BytesIO, int, int]:
    """Resize screenshot to thumbnail; return (BytesIO_png, width_px, height_px)."""
    with PILImage.open(screenshot_path) as img:
        img = img.convert("RGB")
        img.thumbnail((THUMB_W, THUMB_H), PILImage.LANCZOS)
        w, h = img.size
        buf = io.BytesIO()
        img.save(buf, format="PNG", optimize=True)
        buf.seek(0)
    return buf, w, h


def embed_screenshots_in_input(
    excel_path: str,
    df: pd.DataFrame,
    results: list[dict],
) -> None:
    """Open the input Excel file and append a screenshot column in-place."""

    THIN   = Side(style="thin", color="B8CCE4")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

    HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    OK_FILL   = PatternFill("solid", fgColor="E2EFDA")
    ERR_FILL  = PatternFill("solid", fgColor="FCE4D6")
    SKIP_FILL = PatternFill("solid", fgColor="FFF2CC")

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Detect whether the hint row (Vietnamese labels) is present at row 2
    hint_present   = str(ws.cell(2, 1).value or "").strip().startswith("MST")
    data_start_row = 3 if hint_present else 2   # first Excel row that holds actual data

    # New screenshot column = one past the current last column
    ss_col        = ws.max_column + 1
    ss_col_letter = get_column_letter(ss_col)

    # ── Header cell (row 1) ───────────────────────────────────────────────────
    h1 = ws.cell(row=1, column=ss_col, value="ket_qua")
    h1.fill      = HDR_FILL
    h1.font      = HDR_FONT
    h1.border    = BORDER
    h1.alignment = CENTER
    ws.row_dimensions[1].height = max(ws.row_dimensions[1].height or 0, 28)

    # ── Hint-row label (row 2, if present) ───────────────────────────────────
    if hint_present:
        h2 = ws.cell(row=2, column=ss_col, value="Kết quả")
        h2.fill      = PatternFill("solid", fgColor="BDD7EE")
        h2.font      = Font(bold=True, color="1F4E79", size=10, name="Calibri", italic=True)
        h2.border    = BORDER
        h2.alignment = CENTER

    # ── Screenshot column width ───────────────────────────────────────────────
    ws.column_dimensions[ss_col_letter].width = THUMB_W // 7 + 2

    # ── Build row_num → result lookup ─────────────────────────────────────────
    result_by_row = {r["row"]: r for r in results}

    # ── Embed one image per data row ──────────────────────────────────────────
    for df_idx in range(len(df)):
        row_num   = df_idx + 1                      # 1-based, matches results list
        excel_row = data_start_row + df_idx          # actual Excel row
        result    = result_by_row.get(row_num, {})
        status    = result.get("status", "skip")

        cell           = ws.cell(row=excel_row, column=ss_col)
        cell.border    = BORDER
        cell.alignment = CENTER

        if status == "ok" and result.get("file"):
            try:
                buf, w, h = _make_thumb(result["file"])
                xl_img = XLImage(buf)
                
                # Keep high-res image data, but scale down visual display size in Excel
                DISPLAY_W = 800
                scale = min(1.0, DISPLAY_W / w) if w > 0 else 1.0
                display_w = int(w * scale)
                display_h = int(h * scale)
                
                xl_img.width  = display_w
                xl_img.height = display_h
                ws.add_image(xl_img, f"{ss_col_letter}{excel_row}")
                
                ws.row_dimensions[excel_row].height = display_h * 0.75 + 10
                ws.column_dimensions[ss_col_letter].width = display_w / 7.5
            except Exception as exc:
                cell.value = f"Image error: {exc}"
                cell.fill  = ERR_FILL
        elif status == "error":
            cell.value = f"LỖI\n{result.get('error', '')}"
            cell.fill  = ERR_FILL
            cell.font  = Font(color="C00000", size=10, name="Calibri")
            ws.row_dimensions[excel_row].height = 40
        else:
            cell.value = "BỎ QUA"
            cell.fill  = SKIP_FILL
            cell.font  = Font(color="7F6000", size=10, name="Calibri")
            ws.row_dimensions[excel_row].height = 22

    wb.save(excel_path)
    print(f"\nScreenshots embedded → {excel_path}")


# ── Async processing (one browser, N concurrent tabs) ─────────────────────────

import asyncio
from playwright.async_api import async_playwright

DEFAULT_WORKERS = 1


async def _fill_form_async(page, row: dict, auto_solve: bool, captcha_fn=None) -> None:
    """Async version of fill_form — works with async Playwright page."""
    await page.wait_for_selector(".ant-form", timeout=30000)

    # Dismiss popup modal
    try:
        close_btn = page.locator(
            ".ant-modal-close, "
            ".ant-modal-wrap button.ant-modal-close, "
            "button[aria-label='Close']"
        ).first
        if await close_btn.count() and await close_btn.is_visible():
            await close_btn.click(timeout=3000)
            await page.wait_for_timeout(500)
    except Exception:
        pass
    try:
        modal = page.locator(".ant-modal-wrap").first
        if await modal.count() and await modal.is_visible():
            await page.keyboard.press("Escape")
            await page.wait_for_timeout(500)
    except Exception:
        pass

    # MST
    mst = str(row.get("nbmst", "")).strip()
    if mst:
        inp = page.locator(
            'input[id*="nbmst"], input[placeholder*="MST"]'
        ).or_(
            page.locator(".ant-form-item").filter(has_text="MST người bán").locator("input")
        ).first
        await inp.click(timeout=3000)
        await inp.press("Meta+a")
        await inp.press("Backspace")
        await inp.fill(mst)

    # Invoice type
    lhdon = str(row.get("lhdon", "GTGT"))
    lhdon_upper = lhdon.upper().strip() if lhdon else "GTGT"
    label_text = INVOICE_TYPE_MAP.get(lhdon_upper, INVOICE_TYPE_MAP["GTGT"])
    
    # Check if we even need to click
    current_val_loc = page.locator('#lhdon .ant-select-selection-selected-value, #lhdon .ant-select-selection-item').first
    try:
        current_val = await current_val_loc.inner_text(timeout=1000)
        already_selected = (label_text.lower() in current_val.lower())
    except:
        already_selected = False

    if not already_selected:
        trigger = page.locator('#lhdon, .ant-form-item:has-text("Loại hóa đơn") .ant-select').first
        await trigger.click(timeout=5000)
        await page.wait_for_selector(".ant-select-dropdown-menu-item:visible, .ant-select-item-option:visible", timeout=5000)
        
        option = page.locator(
            f'.ant-select-dropdown-menu-item[title*="{label_text}"], '
            f'.ant-select-dropdown-menu-item:has-text("{label_text}"), '
            f'.ant-select-item-option[title*="{label_text}"], '
            f'.ant-select-item-option:has-text("{label_text}")'
        ).first
        
        if await option.count():
            await option.click()
        else:
            await page.locator(".ant-select-dropdown-menu-item, .ant-select-item-option").first.click()

    # khhdon
    khhdon = str(row.get("khhdon", "")).strip()
    if khhdon:
        inp = page.locator(
            'input[id*="khhdon"], input[placeholder*="Ký hiệu"]'
        ).or_(
            page.locator(".ant-form-item").filter(has_text="Ký hiệu hóa đơn").locator("input")
        ).first
        await inp.click(timeout=3000)
        await inp.press("Meta+a")
        await inp.press("Backspace")
        await inp.fill(khhdon)

    # shdon
    shdon = str(row.get("shdon", "")).strip()
    if shdon:
        inp = page.locator(
            'input[id*="shdon"], input[placeholder*="Số hóa đơn"]'
        ).or_(
            page.locator(".ant-form-item").filter(has_text="Số hóa đơn").locator("input")
        ).first
        await inp.click(timeout=3000)
        await inp.press("Meta+a")
        await inp.press("Backspace")
        await inp.fill(shdon)

    # Number fields
    for field_id, label, key in [
        ("tgtthue", "Tổng tiền thuế", "tgtthue"),
        ("tgtttbso", "Tổng tiền thanh toán", "tgtttbso"),
    ]:
        value = row.get(key)
        if value is None or str(value).strip() in ("", "nan"):
            continue
        raw = str(value).replace(",", "").replace(" ", "").strip()
        try:
            raw = str(int(float(raw)))
        except ValueError:
            continue
        inp = page.locator(f"input#{field_id}").first
        if not await inp.count():
            continue
        await inp.fill(raw)
        await inp.press("Tab")

    # CAPTCHA — wait for the image to fully load first
    captcha_img = page.locator(
        "img[src*='captcha'], img[src*='Captcha'], "
        ".ant-form-item:has-text('Mã captcha') img, "
        "img[alt*='captcha']"
    ).first
    try:
        await captcha_img.wait_for(state="visible", timeout=10000)
        # Force refresh to get a crisp new captcha
        await captcha_img.click()
        await asyncio.sleep(0.4) # Tightened from 0.7s
    except Exception:
        pass

    captcha_answer = ""
    # Try up to 2 mini-attempts within the same page to get a valid 6-char answer
    for img_attempt in range(2):
        if captcha_fn:
            captcha_answer = await captcha_fn(page, auto_solve)
        else:
            captcha_answer = get_captcha_answer(page, auto_solve)
        
        # Tax portal captchas are always 6 chars
        if captcha_answer and len(captcha_answer) == 6:
            break
        
        # If still not 6 chars, click to refresh image and loop
        try:
            await captcha_img.click()
            await asyncio.sleep(0.6)
        except:
            break

    if not captcha_answer:
        raise ValueError("Không thể lấy được mã Captcha. Đang thử lại...")

    captcha_inp = page.locator(
        'input[id*="captcha"], input[placeholder*="captcha"], '
        'input[placeholder*="Captcha"], input[placeholder*="mã"]'
    ).or_(
        page.locator(".ant-form-item").filter(has_text="Nhập mã captcha").locator("input")
    ).first
    
    await captcha_inp.click(timeout=3000)
    await asyncio.sleep(0.1)
    await captcha_inp.fill(captcha_answer) # Immediate fill (Turbo mode)
    # No more await captcha_inp.type(captcha_answer, delay=30)

    # Direct Submit via script (Faster than simulated click)
    try:
        await page.evaluate("document.querySelector('button[type=\"submit\"]').click()")
    except:
        await page.locator('button[type="submit"]:has-text("Tìm kiếm")').first.click()
    
    # DO NOT wait for networkidle (it hangs for 30s). Wait for any loading spinners to vanish.
    await asyncio.sleep(0.5)
    
    # Check for Server 500 error notifications (Ant Design toast)
    try:
        error_toast = page.locator(".ant-notification-notice-message, .ant-message-error").first
        if await error_toast.is_visible(timeout=500):
            err_text = await error_toast.inner_text()
            if "500" in err_text or "failed" in err_text.lower():
                print(f"    [WARNING] Server Thuế báo lỗi 500. Đang tải lại trang...")
                await page.goto(LOOKUP_URL, timeout=30000)
                raise ValueError("Server Thuế lỗi (500). Đang thử lại...")
    except:
        pass
    
    await asyncio.sleep(0.5)

    # Check for CAPTCHA error popups and Network Error toast
    err_text = page.locator('.ant-message-notice-content, .ant-modal-confirm-content, .ant-notification-notice-content, .ant-alert-error')
    texts = await err_text.all_inner_texts()
    for text in texts:
        t = text.lower()
        if "network error" in t or "network" in t and "error" in t:
            raise ConnectionError(f"Network Error from server: {text.strip()}")
        if "captcha" in t or "xác thực" in t or "không đúng" in t or "sai" in t:
            raise ValueError(f"CAPTCHA Error or System Error: {text.strip()}")


async def _process_row_async(
    context, row: dict, row_num: int, total: int,
    auto_solve: bool, captcha_fn=None,
) -> dict:
    """Process one row using the shared context (new tab per row)."""
    last_error = ""

    USER_AGENTS = [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:122.0) Gecko/20100101 Firefox/122.0",
        "Mozilla/5.0 (ApplePC; MacOs) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    ]

    for attempt in range(1, MAX_CAPTCHA_RETRIES + 1):
        if attempt > 1:
            backoff = 2.0 + random.random() * 2
            print(f"  [Row {row_num}] Retrying ({attempt}/{MAX_CAPTCHA_RETRIES}) after {backoff:.1f}s…")
            await asyncio.sleep(backoff)
            
        # Create a fresh isolated context with random identity for each attempt
        ua = random.choice(USER_AGENTS)
        page_context = await context.browser.new_context(
            user_agent=ua,
            viewport={"width": 1400 + random.randint(-100, 100), "height": 900},
            ignore_https_errors=True
        )
        page = await page_context.new_page()
        try:
            await page.goto(LOOKUP_URL, wait_until="domcontentloaded", timeout=60000)
            await _fill_form_async(page, row, auto_solve, captcha_fn=captcha_fn)

            kh_val = str(row.get("khhdon", "")).strip() or "none"
            sh_val = str(row.get("shdon", "")).strip() or "none"
            kh_val = "".join(c for c in kh_val if c.isalnum())
            sh_val = "".join(c for c in sh_val if c.isalnum())
            out = OUTPUT_DIR / f"Row{row_num:04d}_{kh_val}_{sh_val}.png"
            
            # Screenshot OK
            await page.screenshot(path=str(out), full_page=True)
            print(f"  [Row {row_num}] -> Screenshot OK")
            return {"row": row_num, "status": "ok", "file": str(out)}

        except Exception as exc:
            last_error = str(exc)
            print(f"  [Row {row_num}] [ERROR] {last_error}")
            if "network" in last_error.lower() or "timeout" in last_error.lower():
                await asyncio.sleep(2 + random.random() * 3)
        finally:
            await page.close()
            await page_context.close()

    return {"row": row_num, "status": "error", "error": last_error}


async def process_rows_async(
    excel_path: str, auto_solve: bool, headless: bool,
    captcha_fn=None, workers: int = DEFAULT_WORKERS,
    progress_callback=None
) -> str:
    """Async entry point: one browser, N concurrent tabs."""
    OUTPUT_DIR.mkdir(exist_ok=True)

    print(f"\nReading: {excel_path}")
    df = pd.read_excel(excel_path, header=0)
    df.columns = [c.lower().strip() for c in df.columns]
    if len(df) > 0 and str(df.iloc[0, 0]).strip().startswith("MST"):
        df = df.iloc[1:].reset_index(drop=True)
    df = df.dropna(how="all").reset_index(drop=True)
    print(f"Found {len(df)} row(s). Columns: {list(df.columns)}")

    MAX_CONCURRENT = 10
    print(f"Processing {len(df)} row(s) with {MAX_CONCURRENT} 'Turbo' persistent tabs\n")
    
    start_time = time.time()
    results: list[dict] = []
    queue = asyncio.Queue()
    for idx, row_series in df.iterrows():
        queue.put_nowait((idx + 1, row_series.to_dict()))

    completed_count = 0

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=headless, args=["--ignore-certificate-errors"])
        
        async def persistent_worker(worker_id):
            nonlocal completed_count
            await asyncio.sleep(worker_id * 0.05) # Even faster stagger
            
            context = await browser.new_context(viewport={"width": 1100, "height": 700}, ignore_https_errors=True)
            page = await context.new_page()

            # TURBO TECHNIQUE 1: Aggressive Blocking (Block fonts, media and non-essential scripts)
            async def block_assets(route):
                url = route.request.url.lower()
                if route.request.resource_type in ["font", "media"]:
                    await route.abort()
                elif "google" in url or "map" in url or "analytics" in url:
                    await route.abort()
                else:
                    await route.continue_()
            await page.route("**/*", block_assets)

            # TURBO TECHNIQUE 2: Disable CSS animations & Hide Popups/Toasts (Clean screenshots)
            await page.add_init_script("""
                const style = document.createElement('style');
                style.innerHTML = `
                    *, *::before, *::after {
                        transition: none !important;
                        animation: none !important;
                    }
                    /* Tàng hình Popups & Notifications */
                    .ant-notification, .ant-modal, .ant-message, .ant-modal-mask {
                        display: none !important;
                        visibility: hidden !important;
                        opacity: 0 !important;
                        pointer-events: none !important;
                    }
                `;
                document.head.appendChild(style);
            """)
            # Initial load
            try: await page.goto(LOOKUP_URL, wait_until="domcontentloaded", timeout=60000)
            except: pass

            while True:
                try:
                    # Get next task from queue safely
                    if queue.empty(): break
                    row_num, row = queue.get_nowait()
                except asyncio.QueueEmpty:
                    break
                except:
                    break

                # Validation
                required = {k: str(row.get(k, "")).strip() for k in ("nbmst", "khhdon", "shdon", "tgtttbso")}
                missing = [k for k, v in required.items() if not v or v == "nan"]
                if missing:
                    print(f"[{row_num}/{len(df)}] SKIP — missing: {missing}")
                    results.append({"row": row_num, "status": "skip", "error": f"missing {missing}"})
                    completed_count += 1
                    if progress_callback: await progress_callback(completed_count, len(df))
                    queue.task_done()
                    continue
                
                print(f"[{row_num}/{len(df)}] Tab {worker_id} processing: shdon={row.get('shdon')}")
                
                success = False
                last_error = ""
                for attempt in range(1, MAX_CAPTCHA_RETRIES + 1):
                    try:
                        if not page.url.startswith("http"):
                            await page.goto(LOOKUP_URL, timeout=30000)
                        
                        await _fill_form_async(page, row, auto_solve, captcha_fn=captcha_fn)
                        
                        kh_val = "".join(c for c in str(row.get("khhdon","")) if c.isalnum())
                        sh_val = "".join(c for c in str(row.get("shdon","")) if c.isalnum())
                        out = OUTPUT_DIR / f"Row{row_num:04d}_{kh_val}_{sh_val}.jpg"
                        
                        # Turbo Photo: JPEG is faster to encode than PNG
                        await page.screenshot(path=str(out), type="jpeg", quality=80)
                        results.append({"row": row_num, "status": "ok", "file": str(out)})
                        success = True
                        break
                    except Exception as exc:
                        last_error = str(exc)
                        if "network" in last_error.lower() or "timeout" in last_error.lower():
                            await asyncio.sleep(1)
                            try: await page.goto(LOOKUP_URL, timeout=30000)
                            except: pass
                        else:
                            await asyncio.sleep(0.5)

                if not success:
                    results.append({"row": row_num, "status": "error", "error": last_error})
                
                completed_count += 1
                if progress_callback: await progress_callback(completed_count, len(df))
                queue.task_done()
            await context.close()

        worker_tasks = [asyncio.create_task(persistent_worker(i)) for i in range(MAX_CONCURRENT)]
        await asyncio.gather(*worker_tasks)
        await browser.close()

    results.sort(key=lambda r: r["row"])

    print("\n" + "=" * 50)
    end_time = time.time()
    duration = end_time - start_time
    minutes = int(duration // 60)
    seconds = int(duration % 60)
    
    ok_count = sum(1 for r in results if r["status"] == "ok")
    print(f"Done — {ok_count}/{len(results)} succeeded!")
    print(f"⏱️ Tổng thời gian xử lý: {minutes}p {seconds}s")
    for r in results:
        sym    = {"ok": "✓", "error": "✗", "skip": "○"}.get(r["status"], "?")
        detail = r.get("file") or r.get("error", "")
        print(f"  {sym} Row {r['row']}: {detail}")

    import shutil
    import os
    zip_path = str(Path("output_images").absolute())
    shutil.make_archive(zip_path, 'zip', str(OUTPUT_DIR))
    
    return zip_path + ".zip"


def process_rows(
    excel_path: str, auto_solve: bool, headless: bool,
    captcha_fn=None, workers: int = DEFAULT_WORKERS,
) -> str:
    """Sync wrapper — used by CLI. Runs the async version internally."""
    return asyncio.run(process_rows_async(
        excel_path, auto_solve, headless, captcha_fn, workers,
    ))


# ── CLI ───────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Auto-fill hoadondientu.gdt.gov.vn/lookup from Excel, save results with screenshots.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("excel", help="Input .xlsx file")
    parser.add_argument(
        "--auto-captcha", action="store_true",
        help="Use Claude AI to solve CAPTCHA (requires ANTHROPIC_API_KEY)",
    )
    parser.add_argument(
        "--headless", action="store_true",
        help="Run browser headlessly (default: visible window)",
    )
    args = parser.parse_args()

    if not os.path.isfile(args.excel):
        print(f"Error: file not found — {args.excel}", file=sys.stderr)
        sys.exit(1)

    if args.auto_captcha and not os.getenv("ANTHROPIC_API_KEY"):
        print("Warning: ANTHROPIC_API_KEY not set — falling back to manual CAPTCHA.")

    process_rows(args.excel, args.auto_captcha, args.headless)


if __name__ == "__main__":
    main()
