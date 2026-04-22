#!/usr/bin/env python3
"""Generate invoices_sample.xlsx — a ready-to-use template for tool.py."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "invoices"

# ── Column definitions ────────────────────────────────────────────────────────
# (field_name, column_header_vn, width, comment_text)
COLUMNS = [
    ("nbmst",    "MST người bán",         18,
     "Mã số thuế người bán.\nBắt buộc. 10 hoặc 14 chữ số.\nVí dụ: 0101283888"),
    ("lhdon",    "Loại hóa đơn",          16,
     "Loại hóa đơn. Mặc định: GTGT\nGTGT  – Hóa đơn GTGT\n"
     "HDBH  – Hóa đơn bán hàng\nHBBTSC – Hóa đơn bán tài sản công\n"
     "HDDTQG – Hóa đơn dự trữ quốc gia\nHDK   – Hóa đơn khác\nCT    – Chứng từ"),
    ("khhdon",   "Ký hiệu hóa đơn",       18,
     "Ký hiệu hóa đơn.\nBắt buộc.\nVí dụ: 1C24TLL, K24TWW"),
    ("shdon",    "Số hóa đơn",            14,
     "Số hóa đơn.\nBắt buộc. Chỉ nhập chữ số.\nVí dụ: 1, 25, 100003"),
    ("tgtthue",  "Tổng tiền thuế",        20,
     "Tổng tiền thuế (VNĐ).\nKhông bắt buộc.\nNhập 0 nếu không có thuế."),
    ("tgtttbso", "Tổng tiền thanh toán",  22,
     "Tổng tiền thanh toán (VNĐ).\nBắt buộc.\nVí dụ: 2420000"),
]

# ── Styles ────────────────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")   # dark blue
SUBHEAD_FILL  = PatternFill("solid", fgColor="BDD7EE")   # light blue
EVEN_FILL     = PatternFill("solid", fgColor="F2F7FC")
ODD_FILL      = PatternFill("solid", fgColor="FFFFFF")

HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
SUBHEAD_FONT  = Font(bold=True, color="1F4E79", size=10, name="Calibri", italic=True)
DATA_FONT     = Font(size=11, name="Calibri")

THIN = Side(style="thin", color="B8CCE4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")

def style_header(cell, text, fill=HEADER_FILL, font=HEADER_FONT):
    cell.value = text
    cell.fill = fill
    cell.font = font
    cell.alignment = CENTER
    cell.border = BORDER

def style_data(cell, value, row_idx):
    cell.value = value
    cell.fill = EVEN_FILL if row_idx % 2 == 0 else ODD_FILL
    cell.font = DATA_FONT
    cell.alignment = CENTER
    cell.border = BORDER

# ── Row 1: machine-readable field names (used by tool.py) ────────────────────
for col, (field, _, width, comment_text) in enumerate(COLUMNS, start=1):
    cell = ws.cell(row=1, column=col)
    style_header(cell, field)
    cmt = Comment(comment_text, "tool.py")
    cmt.width = 260
    cmt.height = 120
    cell.comment = cmt
    ws.column_dimensions[get_column_letter(col)].width = width

ws.row_dimensions[1].height = 28

# ── Row 2: Vietnamese labels (visual hint, light blue) ───────────────────────
for col, (_, label, _, _) in enumerate(COLUMNS, start=1):
    cell = ws.cell(row=2, column=col)
    style_header(cell, label, fill=SUBHEAD_FILL, font=SUBHEAD_FONT)

ws.row_dimensions[2].height = 20

# ── Sample data rows (start at row 3) ────────────────────────────────────────
samples = [
    # nbmst         lhdon   khhdon      shdon    tgtthue   tgtttbso
    ("0312650437",  "GTGT",  "C26MAB",  "363302",  "",      186000),
    ("",            "",      "",         "",        "",       ""),   # blank row for user to fill
    ("",            "",      "",         "",        "",       ""),
]

for r_offset, row_data in enumerate(samples):
    row_idx = r_offset + 3
    for col, val in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col)
        style_data(cell, val if val != "" else None, r_offset)

# ── Freeze rows 1-2 so headers stay visible while scrolling ──────────────────
ws.freeze_panes = "A3"

# ── Data validation: lhdon dropdown ──────────────────────────────────────────
from openpyxl.worksheet.datavalidation import DataValidation
dv = DataValidation(
    type="list",
    formula1='"GTGT,HDBH,HBBTSC,HDDTQG,HDK,CT"',
    allow_blank=True,
    showDropDown=False,
    showErrorMessage=True,
    errorTitle="Giá trị không hợp lệ",
    error="Chọn một trong: GTGT, HDBH, HBBTSC, HDDTQG, HDK, CT",
)
dv.sqref = "B3:B1000"
ws.add_data_validation(dv)

# ── Number format for money columns ──────────────────────────────────────────
for row_idx in range(3, 3 + len(samples)):
    for col in (5, 6):   # tgtthue, tgtttbso
        ws.cell(row=row_idx, column=col).number_format = '#,##0'

# ── Note in sheet tab color ───────────────────────────────────────────────────
ws.sheet_properties.tabColor = "1F4E79"

# ── Save ──────────────────────────────────────────────────────────────────────
out = "invoices_sample.xlsx"
wb.save(out)
print(f"Created: {out}")
print()
print("HOW TO USE:")
print("  1. Edit this file — fill rows from row 3 downward")
print("  2. Row 1 = machine field names (DO NOT change)")
print("  3. Row 2 = Vietnamese labels (for reference only)")
print()
print("  python3 tool.py invoices_sample.xlsx")
print("  python3 tool.py invoices_sample.xlsx --auto-captcha   # AI CAPTCHA")
